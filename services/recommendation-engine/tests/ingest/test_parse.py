"""Parsers: csv/xlsx bytes -> the {canonical_name: rows} shape the validator consumes."""
import io

import openpyxl

from trax_io_reco.ingest.parse import parse_csv, parse_uploads, parse_xlsx


def _single_sheet_xlsx(headers, *rows) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"  # deliberately NOT a canonical name — per-file parse ignores title
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_csv_lowercases_headers():
    data = b"Part_Number,On_Hand\nP1,5\n"
    rows = parse_csv("stock", data)
    assert rows == [{"part_number": "P1", "on_hand": "5"}]


def test_parse_uploads_merges_files():
    files = {"parts": b"part_number\nP1\n", "stock": b"part_number,location_code,on_hand\nP1,MIA,5\n"}
    parsed = parse_uploads(files)
    assert set(parsed) == {"parts", "stock"}
    assert parsed["stock"][0]["location_code"] == "MIA"


def test_parse_xlsx_sheets(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "parts"
    ws.append(["part_number"])
    ws.append(["P1"])
    ws2 = wb.create_sheet("stock")
    ws2.append(["part_number", "location_code", "on_hand"])
    ws2.append(["P1", "MIA", "5"])
    p = tmp_path / "u.xlsx"
    wb.save(p)
    parsed = parse_xlsx(p.read_bytes())
    assert parsed["parts"] == [{"part_number": "P1"}]
    assert parsed["stock"][0]["on_hand"] == "5"


def test_parse_uploads_content_sniffs_per_file_xlsx():
    # The live flow stores each canonical file under an EXTENSION-LESS key, so format
    # detection must be by content. A per-file .xlsx (single sheet, non-canonical title)
    # dropped into the "stock" slot must parse as the stock canonical file, not CSV.
    xlsx_bytes = _single_sheet_xlsx(
        ["part_number", "location_code", "on_hand"], ["P1", "MIA", "5"]
    )
    assert xlsx_bytes[:4] == b"PK\x03\x04"  # sanity: real .xlsx ZIP magic
    files = {"parts": b"part_number\nP1\n", "stock": xlsx_bytes}
    parsed = parse_uploads(files)
    assert parsed["stock"] == [{"part_number": "P1", "location_code": "MIA", "on_hand": "5"}]
    assert parsed["parts"] == [{"part_number": "P1"}]  # CSV sibling still CSV-parsed


def test_parse_uploads_mixed_csv_and_xlsx_same_batch():
    # A batch may mix formats file-by-file; each is sniffed independently.
    files = {
        "parts": _single_sheet_xlsx(["part_number", "criticality"], ["P1", "AOG"]),
        "stock": b"part_number,location_code,on_hand\nP1,JFK,3\n",
    }
    parsed = parse_uploads(files)
    assert parsed["parts"] == [{"part_number": "P1", "criticality": "AOG"}]
    assert parsed["stock"][0]["location_code"] == "JFK"
