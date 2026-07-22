"""Parsers: raw upload bytes -> the ``{canonical_name: rows}`` shape ``validate()`` and
Task 3's mapper consume. Two upload shapes are supported per canonical file: a **CSV**, or
a **single-sheet .xlsx** (the per-file dropzone model — each of the six canonical files is
uploaded on its own, as either format). Format is detected by *content* (the .xlsx ZIP
magic bytes), NOT by filename/extension — the live flow stores each upload under an
extension-less ``{tenant_uuid}/{batch_id}/{canonical_name}`` Storage key, so a suffix check
could never fire. ``parse_xlsx`` additionally supports one multi-sheet workbook keyed by
canonical sheet titles (an alternate bulk shape, not used by the per-file live flow).
Headers are lowercased+stripped; every cell value is kept as a stripped string (the engine
extract format is string-native — see ``extract_loader``'s coercion helpers) so no
numeric/date parsing happens here.
"""

from __future__ import annotations

import csv
import io

import openpyxl

from trax_io_reco.ingest.canonical import CANONICAL_FILES

# .xlsx (like every OOXML/ZIP container) begins with the local-file-header magic
# ``PK\x03\x04``. Content-sniffing is the only reliable format signal because minted
# Storage paths carry no extension (see module docstring).
_XLSX_MAGIC = b"PK\x03\x04"


def _is_xlsx(data: bytes) -> bool:
    return data[:4] == _XLSX_MAGIC


def _clean_header(name: object) -> str:
    return str(name if name is not None else "").strip().lower()


def _clean_value(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_csv(name: str, data: bytes) -> list[dict]:
    """Parse one canonical file's CSV bytes via stdlib ``csv.DictReader``.

    ``name`` identifies the canonical file this CSV belongs to (used by callers for
    routing/errors); the parse itself is header-driven, not name-driven.
    """
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    headers = [_clean_header(h) for h in reader.fieldnames]
    rows: list[dict] = []
    for raw_row in reader:
        rows.append(
            {
                header: _clean_value(raw_row.get(original))
                for header, original in zip(headers, reader.fieldnames, strict=True)
            }
        )
    return rows


def parse_xlsx(data: bytes) -> dict[str, list[dict]]:
    """Parse one workbook where each sheet holds one canonical file's rows. Sheet titles
    are lowercased+stripped; sheets whose (lowercased) title isn't a known canonical file
    name are ignored."""
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    parsed: dict[str, list[dict]] = {}
    for sheet in workbook.worksheets:
        title = _clean_header(sheet.title)
        if title not in CANONICAL_FILES:
            continue
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            parsed[title] = []
            continue
        headers = [_clean_header(h) for h in header_row]
        rows: list[dict] = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            row = {
                header: _clean_value(value)
                for header, value in zip(headers, values, strict=False)
                if header
            }
            rows.append(row)
        parsed[title] = rows
    return parsed


def parse_xlsx_file(name: str, data: bytes) -> list[dict]:
    """Parse ONE canonical file supplied as a single-sheet .xlsx workbook — the per-file
    upload model (each dropzone accepts a CSV or a one-sheet workbook). Reads the FIRST
    worksheet regardless of its title, unlike ``parse_xlsx`` (which routes a multi-sheet
    workbook by canonical sheet title). ``name`` identifies the canonical file for callers;
    the parse itself is header-driven."""
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheets = workbook.worksheets
    if not sheets:
        return []
    rows_iter = sheets[0].iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return []
    headers = [_clean_header(h) for h in header_row]
    rows: list[dict] = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        row = {
            header: _clean_value(value)
            for header, value in zip(headers, values, strict=False)
            if header
        }
        rows.append(row)
    return rows


def parse_uploads(files: dict[str, bytes], *, xlsx: bytes | None = None) -> dict[str, list[dict]]:
    """Merge per-file uploads (each CSV *or* single-sheet .xlsx, detected by content) and/or
    one multi-sheet workbook into the ``{canonical_name: rows}`` dict ``validate()``/the
    mapper consume. Files are keyed by canonical name (case-insensitive); unknown keys are
    ignored. A per-file upload wins over an ``xlsx`` workbook sheet of the same name."""
    parsed: dict[str, list[dict]] = {}
    if xlsx is not None:
        parsed.update(parse_xlsx(xlsx))
    for name, data in files.items():
        key = _clean_header(name)
        if key not in CANONICAL_FILES:
            continue
        parsed[key] = parse_xlsx_file(key, data) if _is_xlsx(data) else parse_csv(key, data)
    return parsed
